"""CLI 入口（对应 VBA 生产工作簿的两个按钮：仅运行校验 / 开始分配，需求 §6.7.3）。

用法：
    python -m returned_inventory validate <工作簿.xlsx>   # Dry Run：仅运行校验
    python -m returned_inventory allocate <工作簿.xlsx>   # Full Run：开始分配

退出码：
    0  成功
    1  其他未预期错误（含工作簿文件不存在）
    2  配置错误（ConfigError）
    3  E12 输入结构异常（InputError）：已中止运行，仅追加运行历史
    4  E99 工程守卫触发（E99Error）：已中止运行，不写输出表

运行成功后会把输出表写回同一个工作簿文件（对应 VBA 直接操作 ThisWorkbook）。
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl

from .config import ConfigError
from .excel_input import InputError
from .guards import E99Error
from .runner import run_full_allocation, run_validation_only

EXIT_OK = 0
EXIT_UNEXPECTED = 1
EXIT_CONFIG_ERROR = 2
EXIT_INPUT_E12 = 3
EXIT_E99 = 4


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="returned_inventory",
        description="退货入库分配系统（VBA 实现的 Python 移植版）",
    )
    subparsers = parser.add_subparsers(dest="command", required=True)
    for name, help_text in (
        ("validate", "仅运行校验（Dry Run）：只检查数据，不执行分配"),
        ("allocate", "开始分配（Full Run）：校验通过后执行回溯分配"),
    ):
        sub = subparsers.add_parser(name, help=help_text)
        sub.add_argument("workbook", help="工作簿 .xlsx 文件路径")
    init_parser = subparsers.add_parser(
        "init", help="新建空白模板工作簿（含全部工作表与表头）"
    )
    init_parser.add_argument("workbook", help="要创建的 .xlsx 文件路径")

    args = parser.parse_args(argv)

    workbook_path = Path(args.workbook)

    if args.command == "init":
        if workbook_path.exists():
            print(f"错误：文件已存在 [{workbook_path}]，为避免覆盖未创建。", file=sys.stderr)
            return EXIT_UNEXPECTED
        create_template_workbook(workbook_path)
        print(f"已创建模板工作簿：{workbook_path}")
        print("请在 输入_退单表 / 输入_质检库存表 粘贴数据后运行 validate 或 allocate。")
        return EXIT_OK

    if not workbook_path.is_file():
        print(f"错误：找不到工作簿文件 [{workbook_path}]。", file=sys.stderr)
        return EXIT_UNEXPECTED

    wb = openpyxl.load_workbook(workbook_path)

    try:
        if args.command == "validate":
            stats = run_validation_only(wb)
        else:
            stats = run_full_allocation(wb)
    except ConfigError as exc:
        print(f"配置读取失败：{exc}\n请修正 输入_配置 后重试。", file=sys.stderr)
        return EXIT_CONFIG_ERROR
    except InputError as exc:
        # E12：结构异常在数据加载阶段中止，仅运行历史已追加一条中止记录
        wb.save(workbook_path)
        print(f"E12 输入结构异常：{exc}\n已中止运行，未生成输出表。", file=sys.stderr)
        return EXIT_INPUT_E12
    except E99Error as exc:
        # E99：工程守卫触发，输出表已清空但不写新结果（对齐 VBA E99Fail 分支）
        wb.save(workbook_path)
        print(
            f"E99 工程守卫触发：{exc}\n已中止运行，请检查库存数据一致性后重试。",
            file=sys.stderr,
        )
        return EXIT_E99

    wb.save(workbook_path)

    if args.command == "validate":
        print(f"干跑完成。\n校验失败物流单号：{stats.validation_fail_count} 个")
    else:
        print(
            f"完整分配完成。\n成功：{stats.alloc_success_count} 个物流单号"
            f"\n失败：{stats.alloc_fail_count} 个物流单号"
        )
    return EXIT_OK


def create_template_workbook(path: Path) -> None:
    """创建空白模板工作簿：8 张工作表 + 表头 + 默认配置示例行。

    输出表只写表头（数据区由每次运行清空重写）；配置表预填全部默认值，
    用户可直接改「值」列。注意 .xlsm 会丢宏，模板固定用 .xlsx。
    """
    from .excel_input import INVENTORY_HEADERS, RETURN_HEADERS
    from .excel_output import RUN_HISTORY_HEADERS
    from .models import (
        DEFAULT_DEBUG_LOG_LEVEL,
        DEFAULT_DETAILED_LOG_LIMIT,
        DEFAULT_LOT_MODE,
        DEFAULT_MAX_BACKTRACK_COUNT,
        DEFAULT_NO_EXPIRY_SENTINEL,
    )

    wb = openpyxl.Workbook()
    wb.active.title = "输入_退单表"
    wb["输入_退单表"].append(RETURN_HEADERS)
    wb.create_sheet("输入_质检库存表").append(INVENTORY_HEADERS)

    ws_config = wb.create_sheet("输入_配置")
    ws_config.append(["参数名", "值", "说明"])
    for name, value, note in (
        ("最大回溯次数", DEFAULT_MAX_BACKTRACK_COUNT, "每个(物流单号+SKU)组的回溯上限，超出触发 E10"),
        ("调试日志级别", DEFAULT_DEBUG_LOG_LEVEL, "关闭 / 简版 / 详细"),
        ("详细日志单表上限", DEFAULT_DETAILED_LOG_LIMIT, "仅详细模式生效，超过自动分表"),
        ("批号比较模式", DEFAULT_LOT_MODE, "不敏感（统一大写）/ 敏感（保留原样）"),
        ("无保质期哨兵值", DEFAULT_NO_EXPIRY_SENTINEL, "无保质期商品的占位效期"),
    ):
        ws_config.append([name, value, note])

    for name in ("分配状态汇总表", "成功分配明细表", "数据异常明细表", "调试日志"):
        wb.create_sheet(name)
    wb.create_sheet("运行历史记录表").append(RUN_HISTORY_HEADERS)

    wb.save(path)


if __name__ == "__main__":
    raise SystemExit(main())
