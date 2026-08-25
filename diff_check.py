"""阶段11对拍：逐工作表逐单元格比对 VBA 版与 Python 版输出。

前置：
    1. ./.venv/Scripts/python.exe diff_build_inputs.py   # 生成输入
    2. powershell -ExecutionPolicy Bypass -File ./diff_run_vba.ps1   # VBA 侧运行
    3. 对同一份输入跑 Python CLI allocate（本脚本 --run-python 可代跑）

用法：
    ./.venv/Scripts/python.exe diff_check.py [--run-python]

比对规则：
    - 五张输出表：分配状态汇总表/成功分配明细表/数据异常明细表/调试日志/运行历史记录表
    - 运行历史记录表的「运行时间/校验耗时/分配耗时/总耗时」列豁免比对（两边天然不同）
    - 数值归一化：12.0 与 12 视为一致；None 与空串视为一致
"""

from __future__ import annotations

import argparse
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
DIFF_DIR = ROOT / "data" / "diff"

CASES = ["TC21", "TC24", "TC20", "TC19", "TC36", "TCXB"]

# 已知行为偏差（对拍前已核实，属 VBA 侧缺陷/旧文档叙述，不计为移植 bug）：
# TCXB 命中预检测B 碎片场景（S<T<S+minQtyOther）：Python 按需求正文实现完整判定，
# VBA modSortFilter.bas 只判 supply<forcedDemand 漏检 → 两边输出必然不一致。
EXPECTED_DIVERGENT = {
    "TCXB": "预检测B碎片场景（S<T<S+minQtyOther）：VBA 预检测B 仅判 supply<forcedDemand，漏检碎片（VBA 侧已知缺陷）",
}

COMPARE_SHEETS = ["分配状态汇总表", "成功分配明细表", "数据异常明细表", "调试日志", "运行历史记录表"]
HISTORY_SHEET = "运行历史记录表"
HISTORY_EXEMPT_HEADERS = {"运行时间", "校验耗时（秒）", "分配耗时（秒）", "总耗时（秒）"}
DEBUG_SHEET = "调试日志"


def norm(value) -> str:
    """单元格值归一化为可比对字符串。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d %H:%M:%S")
    return str(value)


def read_sheet(ws) -> list[list[str]]:
    max_row = max(ws.max_row, 1)
    max_col = max(ws.max_column, 1)
    grid = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
        grid.append([norm(v) for v in row])
    # 去掉尾部全空行（两边 max_row 可能因格式残留不同）
    while grid and all(v == "" for v in grid[-1]):
        grid.pop()
    return grid


def debug_split_sheets(wb) -> list[str]:
    return [
        n
        for n in wb.sheetnames
        if n.startswith(DEBUG_SHEET + "_") and n[len(DEBUG_SHEET) + 1 :].isdigit()
    ]


def compare_sheet(vba_ws, py_ws, sheet_name: str) -> list[str]:
    """返回差异描述列表；空列表表示一致。"""
    vba = read_sheet(vba_ws)
    py = read_sheet(py_ws)
    diffs = []

    exempt_cols: set[int] = set()
    if sheet_name == HISTORY_SHEET and vba and py:
        headers = vba[0]
        for idx, h in enumerate(headers):
            if h in HISTORY_EXEMPT_HEADERS:
                exempt_cols.add(idx)

    if vba and py and vba[0] != py[0]:
        diffs.append(f"表头不一致：VBA={vba[0]} PY={py[0]}")

    max_rows = max(len(vba), len(py))
    for r in range(max_rows):
        vrow = vba[r] if r < len(vba) else []
        prow = py[r] if r < len(py) else []
        max_cols = max(len(vrow), len(prow))
        for c in range(max_cols):
            if r == 0 or c in exempt_cols:
                continue
            vv = vrow[c] if c < len(vrow) else ""
            pv = prow[c] if c < len(prow) else ""
            if vv != pv:
                header = vba[0][c] if vba and c < len(vba[0]) else f"列{c + 1}"
                diffs.append(f"行{r + 1} 列[{header}]：VBA=[{vv}] PY=[{pv}]")
    return diffs


def run_python_side() -> None:
    """对每份 <case>_py.xlsx 跑 Python CLI allocate（原地写回）。"""
    for case in CASES:
        path = DIFF_DIR / f"{case}_py.xlsx"
        proc = subprocess.run(
            [sys.executable, "-m", "returned_inventory", "allocate", str(path)],
            cwd=ROOT,
            capture_output=True,
        )
        # Windows 控制台输出为 GBK，统一按字节捕获后容错解码，避免 UnicodeDecodeError
        msg = (proc.stdout or proc.stderr or b"").decode("utf-8", errors="replace").strip()
        print(f"[python] {case} exit={proc.returncode} {msg}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--run-python", action="store_true", help="先跑 Python CLI allocate 再比对")
    args = parser.parse_args()

    if args.run_python:
        run_python_side()

    total_diff = 0
    unexpected_diff = 0
    for case in CASES:
        vba_path = DIFF_DIR / f"{case}_vba.xlsm"
        py_path = DIFF_DIR / f"{case}_py.xlsx"
        if not vba_path.is_file():
            print(f"[{case}] 缺少 VBA 输出 {vba_path.name}，跳过")
            continue
        vba_wb = openpyxl.load_workbook(vba_path)
        py_wb = openpyxl.load_workbook(py_path)

        vba_splits = debug_split_sheets(vba_wb)
        py_splits = debug_split_sheets(py_wb)
        if vba_splits != py_splits:
            print(f"[{case}] 调试日志分表不一致：VBA={vba_splits} PY={py_splits}")
            total_diff += 1

        print(f"[{case}]")
        case_diff = 0
        for sheet in COMPARE_SHEETS:
            if sheet not in vba_wb.sheetnames or sheet not in py_wb.sheetnames:
                print(f"  {sheet}: 缺表 VBA={sheet in vba_wb.sheetnames} PY={sheet in py_wb.sheetnames}")
                case_diff += 1
                continue
            diffs = compare_sheet(vba_wb[sheet], py_wb[sheet], sheet)
            if not diffs:
                rows = len(read_sheet(py_wb[sheet])) - 1
                print(f"  {sheet}: 一致（数据行 {rows}）")
            else:
                case_diff += len(diffs)
                print(f"  {sheet}: 差异 {len(diffs)} 处")
                for d in diffs[:20]:
                    print(f"    {d}")
                if len(diffs) > 20:
                    print(f"    ……其余 {len(diffs) - 20} 处从略")
        if case_diff and case in EXPECTED_DIVERGENT:
            print(f"  >>> 预期分歧：{EXPECTED_DIVERGENT[case]}")
        elif case_diff and case not in EXPECTED_DIVERGENT:
            print("  >>> 异常分歧：需定位分析")
            unexpected_diff += case_diff
        total_diff += case_diff
    print(f"汇总：总差异 {total_diff} 处，其中异常分歧 {unexpected_diff} 处")
    return 1 if unexpected_diff else 0


if __name__ == "__main__":
    raise SystemExit(main())
