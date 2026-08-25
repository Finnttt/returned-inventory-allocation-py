"""文件级集成测试：真实 xlsx 落盘 → 重新加载 → run_full_allocation 全链路。

与 test_e2e_acceptance.py（内存工作簿）的区别：本文件经 openpyxl save/load
真实走一遍文件读写，验证 runner 在落盘工作簿上的行为（对应 VBA 的
RunSingleTest 17 文件集成层，需求 §6.5.3 自动化入口分层）。

三个代表性场景（输入/期望取自根目录 TC-*.md 冻结数据）：
- TC-21（SF3190000000016）：需要回溯的成功场景（回溯 4 步后全成功）；
- TC-24（SF3190000000028）：回溯超限 E10 + 跨 SKU 短路连带回滚（配置 最大回溯次数=10）；
- TC-20（SF3190000000060）：校验错误 E08（数量不一致），不进入分配阶段。

运行后的工作簿同时保存到 python_port/data/ 下，供阶段 11 与 VBA 对拍复用。
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

from returned_inventory.models import (
    DEBUG_LEVEL_SIMPLE,
    STATUS_BATCH_IMPORT,
    STATUS_UNALLOCATED,
)
from returned_inventory.runner import run_full_allocation

from .wb_factory import build_test_workbook

SF16 = "SF3190000000016"  # TC-21
SF28 = "SF3190000000028"  # TC-24
SF60 = "SF3190000000060"  # TC-20

# 阶段 11 对拍复用资产目录（python_port/data/）
DATA_DIR = Path(__file__).resolve().parent.parent / "data"

OUTPUT_SHEETS = ["分配状态汇总表", "成功分配明细表", "数据异常明细表", "调试日志", "运行历史记录表"]


def _data_rows(ws):
    return [
        ["" if cell.value is None else cell.value for cell in row]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row)
    ]


def _run_via_file(wb, tmp_path: Path, asset_name: str):
    """工作簿落盘 → 重新加载 → 完整分配 → 另存一份到 data/ 供对拍，返回 (wb, stats)。"""
    file_path = tmp_path / f"{asset_name}.xlsx"
    wb.save(file_path)

    reloaded = openpyxl.load_workbook(file_path)
    stats = run_full_allocation(reloaded)

    DATA_DIR.mkdir(exist_ok=True)
    reloaded.save(DATA_DIR / f"{asset_name}.xlsx")
    return reloaded, stats


def _assert_five_output_sheets_present(wb):
    for name in OUTPUT_SHEETS:
        assert name in wb.sheetnames, f"缺少输出表 {name}"


# -----------------------------------------------------------------------------
# 场景一（TC-21）：需要回溯的成功场景
# -----------------------------------------------------------------------------


def test_file_tc21_backtrack_success(tmp_path):
    wb = build_test_workbook(
        order_rows=[
            [SF16, "TK10000161", "H000000001", "00001", 12],
            [SF16, "TK10000161", "H000000001", "00002", 5],
            [SF16, "TK10000161", "H000000001", "00003", 5],
            [SF16, "TK10000161", "H000000001", "00004", 5],
            [SF16, "TK10000161", "H000000001", "00005", 5],
            [SF16, "TK10000161", "H000000001", "00006", 5],
            [SF16, "TK10000161", "H000000001", "00007", 5],
            [SF16, "TK10000161", "H000000002", "00008", 3],
            [SF16, "TK10000161", "H000000003", "00009", 1],
            [SF16, "TK10000162", "H000000002", "00001", 2],
        ],
        inventory_rows=[
            [SF16, "H000000001", "ZP", "LA01", "2029/01/01", 8],
            [SF16, "H000000001", "ZP", "LA01", "2029/01/01", 12],
            [SF16, "H000000001", "QC", "LA01", "2029/01/01", 12],
            [SF16, "H000000001", "QC", "LA01", "2029/01/01", 5],
            [SF16, "H000000001", "QC", "LA01", "2029/01/01", 5],
            [SF16, "H000000002", "NG", "LB01", "2029/01/01", 5],
            [SF16, "H000000003", "NG", "LB01", "2029/01/01", 1],
        ],
    )
    wb, stats = _run_via_file(wb, tmp_path, "e2e_SF0016_TC21")

    _assert_five_output_sheets_present(wb)
    assert _data_rows(wb["分配状态汇总表"]) == [
        [SF16, "TK10000161", STATUS_BATCH_IMPORT, ""],
        [SF16, "TK10000162", STATUS_BATCH_IMPORT, ""],
    ]

    # 明细表 10 行（TC-21 §四），抽查关键行：回溯后 00001 改选 QC、00004~00007 走 ZP
    detail = _data_rows(wb["成功分配明细表"])
    assert len(detail) == 10
    assert detail[0] == [SF16, "TK10000161", "H000000001", "00001", 12, "QC", "LA01",
                         "2029/01/01", 12, STATUS_BATCH_IMPORT, STATUS_BATCH_IMPORT]
    assert all(row[5] == "ZP" for row in detail[3:7])  # 00004~00007
    assert detail[8][1] == "TK10000162" and detail[8][8] == 2  # 跨退单号共享库存池

    assert _data_rows(wb["数据异常明细表"]) == []
    assert _data_rows(wb["调试日志"]) == []  # 默认配置关闭

    history = _data_rows(wb["运行历史记录表"])
    assert len(history) == 1
    assert history[0][2] == "Full Run"
    assert history[0][10] == 1 and history[0][11] == 0
    assert history[0][13] == 4 and history[0][14] == 4  # 回溯 4 步后成功

    # 对拍资产已生成且可重新打开
    assert (DATA_DIR / "e2e_SF0016_TC21.xlsx").exists()
    assert "分配状态汇总表" in openpyxl.load_workbook(DATA_DIR / "e2e_SF0016_TC21.xlsx").sheetnames


# -----------------------------------------------------------------------------
# 场景二（TC-24）：回溯超限 E10 + 短路连带回滚（配置敏感：最大回溯次数=10）
# -----------------------------------------------------------------------------


def test_file_tc24_backtrack_limit_e10(tmp_path):
    wb = build_test_workbook(
        order_rows=[
            [SF28, "TK10000281", "H000000001", "00001", 6],
            [SF28, "TK10000282", "H000000001", "00001", 6],
            [SF28, "TK10000282", "H000000001", "00002", 6],
            [SF28, "TK10000282", "H000000001", "00003", 4],
            [SF28, "TK10000282", "H000000001", "00004", 4],
            [SF28, "TK10000282", "H000000001", "00005", 4],
            [SF28, "TK10000282", "H000000001", "00006", 4],
            [SF28, "TK10000282", "H000000001", "00007", 4],
            [SF28, "TK10000282", "H000000001", "00008", 4],
            [SF28, "TK10000282", "H000000002", "00009", 6],
            [SF28, "TK10000282", "H000000002", "00010", 3],
        ],
        inventory_rows=[
            [SF28, "H000000001", "ZP", "LA01", "2029/01/01", 23],
            [SF28, "H000000001", "QC", "LA01", "2029/01/01", 13],
            [SF28, "H000000001", "NG", "LA01", "2029/01/01", 6],
            [SF28, "H000000002", "ZP", "LA01", "2029/01/01", 9],
        ],
        config_rows=[
            ["最大回溯次数", 10, ""],
            ["调试日志级别", DEBUG_LEVEL_SIMPLE, ""],
        ],
    )
    wb, stats = _run_via_file(wb, tmp_path, "e2e_SF0028_TC24")

    _assert_five_output_sheets_present(wb)
    assert _data_rows(wb["分配状态汇总表"]) == [
        [SF28, "TK10000281", STATUS_UNALLOCATED, "E10 - 回溯超限"],
        [SF28, "TK10000282", STATUS_UNALLOCATED, "E10 - 回溯超限"],
    ]
    assert _data_rows(wb["成功分配明细表"]) == []  # 整单回滚
    assert _data_rows(wb["数据异常明细表"]) == []

    # 简版调试日志经文件往返后仍完整：11 行。
    # H1：首失败行（00003）E10/回溯路径穷尽，其余 8 行为同 SKU 连带回滚（E09）；
    # H2：两行跨 SKU 短路（占位码"连带回滚"）。与 VBA BT_FillFailureDebugFields 口径一致；
    # TC-24 文档调试日志表（全行 E10）为旧版叙述，以实现为准。
    debug = _data_rows(wb["调试日志"])
    assert len(debug) == 11
    h1_rows = [row for row in debug if row[1] == "H000000001"]
    assert len(h1_rows) == 9
    assert sum(1 for row in h1_rows if row[17] == "E10") == 1
    assert sum(1 for row in h1_rows if row[17] == "E09") == 8
    assert all(row[17] == "连带回滚" for row in debug if row[1] == "H000000002")

    assert stats.alloc_fail_count == 1
    history = _data_rows(wb["运行历史记录表"])
    assert history[0][12] == "E10:1; 连带回滚:1"
    assert history[0][13] == 11 and history[0][14] == 11
    assert history[0][17] == 10  # 配置快照：最大回溯次数


# -----------------------------------------------------------------------------
# 场景三（TC-20）：校验错误 E08，不进入分配阶段
# -----------------------------------------------------------------------------


def test_file_tc20_e08_validation_error(tmp_path):
    wb = build_test_workbook(
        order_rows=[
            [SF60, "TK10000600", "H000000060", "00001", 5],
            [SF60, "TK10000601", "H000000060", "00001", 3],
        ],
        inventory_rows=[[SF60, "H000000060", "ZP", "LA01", "2029/06/15", 5]],
    )
    wb, stats = _run_via_file(wb, tmp_path, "e2e_SF0060_TC20")

    _assert_five_output_sheets_present(wb)
    assert _data_rows(wb["分配状态汇总表"]) == [
        [SF60, "TK10000600", STATUS_UNALLOCATED, "E08 - 同物流单号+SKU数量不一致"],
        [SF60, "TK10000601", STATUS_UNALLOCATED, "E08 - 同物流单号+SKU数量不一致"],
    ]
    assert _data_rows(wb["成功分配明细表"]) == []
    assert _data_rows(wb["数据异常明细表"]) == []  # E08 跨表汇总级，不进异常明细
    assert _data_rows(wb["调试日志"]) == []

    assert stats.validation_fail_count == 1
    assert stats.alloc_success_count == 0 and stats.alloc_fail_count == 0
    history = _data_rows(wb["运行历史记录表"])
    assert history[0][12] == "E08:1"
    assert history[0][13] == 0 and history[0][14] == 0
