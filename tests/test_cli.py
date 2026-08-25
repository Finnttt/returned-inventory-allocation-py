"""CLI 入口测试（`python -m returned_inventory validate|allocate`，需求 §6.7.3）。

覆盖：两个子命令的成功退出码、E12/配置错误/文件缺失的非零退出码、
运行结果写回工作簿文件。
"""

import openpyxl

from returned_inventory.cli import (
    EXIT_CONFIG_ERROR,
    EXIT_INPUT_E12,
    EXIT_OK,
    EXIT_UNEXPECTED,
    main,
)

from .wb_factory import build_test_workbook, simple_success_workbook


def _save(wb, tmp_path, name="book.xlsx"):
    path = tmp_path / name
    wb.save(path)
    return str(path)


def test_cli_validate_success(tmp_path):
    path = _save(simple_success_workbook(), tmp_path)
    assert main(["validate", path]) == EXIT_OK

    wb = openpyxl.load_workbook(path)
    history = list(wb["运行历史记录表"].iter_rows(min_row=2, values_only=True))
    assert len(history) == 1
    assert history[0][2] == "Dry Run"


def test_cli_allocate_success(tmp_path):
    path = _save(simple_success_workbook(), tmp_path)
    assert main(["allocate", path]) == EXIT_OK

    wb = openpyxl.load_workbook(path)
    detail = list(wb["成功分配明细表"].iter_rows(min_row=2, values_only=True))
    assert len(detail) == 1
    assert detail[0][0] == "SF001"
    history = list(wb["运行历史记录表"].iter_rows(min_row=2, values_only=True))
    assert history[0][2] == "Full Run"


def test_cli_validate_e12_returns_nonzero(tmp_path):
    wb = build_test_workbook(
        order_rows=[["SF001", "WMS001", "SKU-A", "00001", 5]],
        inventory_rows=[["SF001", "SKU-A", "ZP", "LA01", "2029/01/01", 5]],
        order_headers=["物流单号", "WMS退单号", "sku", "行号", "数量"],
    )
    path = _save(wb, tmp_path)

    assert main(["validate", path]) == EXIT_INPUT_E12

    # E12 中止记录已持久化到运行历史
    wb = openpyxl.load_workbook(path)
    history = list(wb["运行历史记录表"].iter_rows(min_row=2, values_only=True))
    assert len(history) == 1
    assert history[0][9] == "[E12-中止]"
    # 不生成输出表
    assert wb["分配状态汇总表"].max_row == 1


def test_cli_config_error_returns_nonzero(tmp_path):
    wb = build_test_workbook(
        order_rows=[["SF001", "WMS001", "SKU-A", "00001", 5]],
        inventory_rows=[["SF001", "SKU-A", "ZP", "LA01", "2029/01/01", 5]],
        config_rows=[["最大回溯次数", "abc", ""]],
    )
    path = _save(wb, tmp_path)

    assert main(["allocate", path]) == EXIT_CONFIG_ERROR


def test_cli_missing_workbook_returns_nonzero(tmp_path):
    assert main(["validate", str(tmp_path / "不存在.xlsx")]) == EXIT_UNEXPECTED


def test_cli_init_creates_runnable_template(tmp_path):
    path = str(tmp_path / "模板.xlsx")
    assert main(["init", path]) == EXIT_OK
    # 已存在的文件不覆盖
    assert main(["init", path]) == EXIT_UNEXPECTED

    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == [
        "输入_退单表",
        "输入_质检库存表",
        "输入_配置",
        "分配状态汇总表",
        "成功分配明细表",
        "数据异常明细表",
        "调试日志",
        "运行历史记录表",
    ]
    # 空模板可直接跑干跑（默认配置 + 空输入，不报错）
    assert main(["validate", path]) == EXIT_OK
