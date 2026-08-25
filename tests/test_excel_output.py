"""M14 Excel 写入单元测试（对应 VBA modExcelOutput.bas，需求 §6.7.4）。

覆盖：清空保留表头、只清输出表不动输入/配置/运行历史、调试日志分表删除与
连续编号重建、关闭模式仅留表头、运行历史追加与运行编号自增、文本格式列。
"""

import openpyxl
import pytest

from returned_inventory.excel_output import (
    DEBUG_HEADERS,
    RUN_HISTORY_HEADERS,
    SUMMARY_HEADERS,
    OutputError,
    append_run_history,
    clear_output_sheets,
    write_debug_log,
    write_sheet,
)
from returned_inventory.models import (
    DEBUG_LEVEL_DETAIL,
    DEBUG_LEVEL_OFF,
    Config,
)


def _rows(values_list):
    return [list(values) for values in values_list]


def _data_rows(ws):
    """读取工作表第 2 行起的数据区（list of list）。"""
    return [
        [cell.value for cell in row]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row)
    ]


# -----------------------------------------------------------------------------
# clear_output_sheets
# -----------------------------------------------------------------------------


def test_clear_keeps_header_and_removes_data():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "分配状态汇总表"
    ws.append(SUMMARY_HEADERS)
    ws.append(["SF001", "WMS001", "批量导入", ""])
    for name in ("成功分配明细表", "数据异常明细表", "调试日志"):
        wb.create_sheet(name)

    clear_output_sheets(wb, Config())

    assert ws.max_row == 1
    assert [cell.value for cell in ws[1]] == SUMMARY_HEADERS


def test_clear_deletes_debug_split_sheets_but_keeps_other_sheets():
    wb = openpyxl.Workbook()
    wb.active.title = "分配状态汇总表"
    for name in ("成功分配明细表", "数据异常明细表", "调试日志"):
        wb.create_sheet(name)
    wb.create_sheet("调试日志_2")
    wb.create_sheet("调试日志_3")
    wb.create_sheet("调试日志_2A")  # 非纯数字后缀，不属于分表
    ws_input = wb.create_sheet("输入_退单表")
    ws_input.append(["物流单号"])
    ws_input.append(["SF001"])
    ws_history = wb.create_sheet("运行历史记录表")
    ws_history.append(RUN_HISTORY_HEADERS)
    ws_history.append([1, "2026/08/22 10:00:00"])

    clear_output_sheets(wb, Config())

    assert "调试日志_2" not in wb.sheetnames
    assert "调试日志_3" not in wb.sheetnames
    assert "调试日志_2A" in wb.sheetnames
    assert "调试日志" in wb.sheetnames
    # 输入表 / 运行历史不在清空范围
    assert ws_input.max_row == 2
    assert ws_history.max_row == 2


def test_clear_protected_sheet_raises_output_error():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "分配状态汇总表"
    ws.protection.sheet = True
    for name in ("成功分配明细表", "数据异常明细表", "调试日志"):
        wb.create_sheet(name)

    with pytest.raises(OutputError, match="受保护"):
        clear_output_sheets(wb, Config())


# -----------------------------------------------------------------------------
# write_sheet
# -----------------------------------------------------------------------------


def test_write_sheet_writes_header_and_data():
    wb = openpyxl.Workbook()
    ws = wb.active
    rows = _rows([["SF001", "WMS001", "批量导入", ""], ["SF002", "WMS002", "无法分配", "E09 - 分配路径穷尽"]])
    write_sheet(ws, rows, SUMMARY_HEADERS)

    assert [cell.value for cell in ws[1]] == SUMMARY_HEADERS
    assert _data_rows(ws) == rows


def test_write_sheet_overwrites_stale_data():
    wb = openpyxl.Workbook()
    ws = wb.active
    write_sheet(ws, _rows([["OLD1"], ["OLD2"], ["OLD3"]]), SUMMARY_HEADERS)
    write_sheet(ws, _rows([["NEW1", "W", "S", "R"]]), SUMMARY_HEADERS)

    assert ws.max_row == 2
    assert _data_rows(ws) == [["NEW1", "W", "S", "R"]]


def test_write_sheet_applies_text_format_to_line_no_column():
    wb = openpyxl.Workbook()
    ws = wb.active
    write_sheet(ws, _rows([["SF001", "WMS001", "SKU-A", "00001"]]), ["物流单号", "WMS退单号", "SKU", "行号"])
    assert ws.cell(row=2, column=4).number_format == "@"
    assert ws.cell(row=2, column=4).value == "00001"


# -----------------------------------------------------------------------------
# write_debug_log
# -----------------------------------------------------------------------------


def _debug_wb():
    wb = openpyxl.Workbook()
    wb.active.title = "调试日志"
    return wb


def test_write_debug_log_splits_with_consecutive_numbering():
    wb = _debug_wb()
    cfg = Config(debug_log_level=DEBUG_LEVEL_DETAIL, detailed_log_limit=3)
    rows = _rows([["R%d" % i] + [""] * 18 for i in range(1, 8)])  # 7 行 → 3+3+1

    write_debug_log(wb, rows, cfg)

    assert wb.sheetnames == ["调试日志", "调试日志_2", "调试日志_3"]
    assert wb["调试日志"].max_row == 1 + 3
    assert wb["调试日志_2"].max_row == 1 + 3
    assert wb["调试日志_3"].max_row == 1 + 1
    assert wb["调试日志_3"].cell(row=2, column=1).value == "R7"
    assert [cell.value for cell in wb["调试日志_2"][1]] == DEBUG_HEADERS


def test_write_debug_log_rerun_removes_stale_splits():
    wb = _debug_wb()
    cfg = Config(debug_log_level=DEBUG_LEVEL_DETAIL, detailed_log_limit=2)
    write_debug_log(wb, _rows([["A"] * 19] * 5), cfg)  # → 3 张分表
    assert "调试日志_3" in wb.sheetnames

    write_debug_log(wb, _rows([["B"] * 19] * 2), cfg)  # 重跑：只剩主表
    assert wb.sheetnames == ["调试日志"]
    assert wb["调试日志"].max_row == 1 + 2


def test_write_debug_log_off_level_keeps_header_only():
    wb = _debug_wb()
    wb.create_sheet("调试日志_2")  # 上次详细模式的残留分表
    cfg = Config(debug_log_level=DEBUG_LEVEL_OFF)

    write_debug_log(wb, _rows([["X"] * 19]), cfg)

    assert wb.sheetnames == ["调试日志"]
    assert wb["调试日志"].max_row == 1
    assert [cell.value for cell in wb["调试日志"][1]] == DEBUG_HEADERS


def test_write_debug_log_non_positive_limit_uses_default():
    wb = _debug_wb()
    cfg = Config(debug_log_level=DEBUG_LEVEL_DETAIL, detailed_log_limit=0)
    write_debug_log(wb, _rows([["A"] * 19] * 3), cfg)
    assert wb.sheetnames == ["调试日志"]
    assert wb["调试日志"].max_row == 1 + 3


# -----------------------------------------------------------------------------
# append_run_history
# -----------------------------------------------------------------------------


def test_append_run_history_auto_increment_run_number():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(RUN_HISTORY_HEADERS)

    append_run_history(ws, ["", "2026/08/22 10:00:00", "Dry Run"] + [0] * 17)
    append_run_history(ws, ["", "2026/08/22 11:00:00", "Full Run"] + [0] * 17)

    assert ws.max_row == 3
    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=3, column=1).value == 2
    assert ws.cell(row=2, column=3).value == "Dry Run"
    assert ws.cell(row=3, column=3).value == "Full Run"


def test_append_run_history_writes_time_and_sentinel_as_text():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(RUN_HISTORY_HEADERS)
    row = [""] * 20
    row[1] = "2026/08/22 10:00:00"
    row[19] = "2099/01/01"

    append_run_history(ws, row)

    assert ws.cell(row=2, column=2).number_format == "@"
    assert ws.cell(row=2, column=20).number_format == "@"


def test_append_run_history_empty_row_is_noop():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(RUN_HISTORY_HEADERS)
    append_run_history(ws, [])
    assert ws.max_row == 1
