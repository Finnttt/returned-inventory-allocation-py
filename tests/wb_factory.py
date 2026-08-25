"""测试共享的工作簿构造辅助（runner / CLI 端到端测试共用）。

构造与生产工作簿同构的内存 openpyxl Workbook：
输入_退单表 / 输入_质检库存表 / 输入_配置 + 四张输出表 + 运行历史记录表。
"""

from __future__ import annotations

import openpyxl

from returned_inventory.excel_input import INVENTORY_HEADERS, RETURN_HEADERS
from returned_inventory.excel_output import RUN_HISTORY_HEADERS

CONFIG_HEADERS = ["参数名", "值", "说明"]

OUTPUT_SHEET_NAMES = ["分配状态汇总表", "成功分配明细表", "数据异常明细表", "调试日志"]


def build_test_workbook(
    order_rows: list[list] | None = None,
    inventory_rows: list[list] | None = None,
    config_rows: list[list] | None = None,
    order_headers: list[str] | None = None,
) -> openpyxl.Workbook:
    """构造测试工作簿。order_headers 可传入错误表头以模拟 E12。"""
    wb = openpyxl.Workbook()

    ws_orders = wb.active
    ws_orders.title = "输入_退单表"
    ws_orders.append(order_headers or RETURN_HEADERS)
    for row in order_rows or []:
        ws_orders.append(row)

    ws_inventory = wb.create_sheet("输入_质检库存表")
    ws_inventory.append(INVENTORY_HEADERS)
    for row in inventory_rows or []:
        ws_inventory.append(row)

    ws_config = wb.create_sheet("输入_配置")
    ws_config.append(CONFIG_HEADERS)
    for row in config_rows or []:
        ws_config.append(row)

    for name in OUTPUT_SHEET_NAMES:
        wb.create_sheet(name)

    ws_history = wb.create_sheet("运行历史记录表")
    ws_history.append(RUN_HISTORY_HEADERS)

    return wb


def simple_success_workbook() -> openpyxl.Workbook:
    """最小成功场景：SF001 / WMS001 / SKU-A / 00001 需求 5，库存 ZP 恰好 5。"""
    return build_test_workbook(
        order_rows=[["SF001", "WMS001", "SKU-A", "00001", 5]],
        inventory_rows=[["SF001", "SKU-A", "ZP", "LA01", "2029/01/01", 5]],
    )
