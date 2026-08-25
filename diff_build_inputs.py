"""阶段11对拍：构造各对拍场景的输入工作簿（Python 侧）与 JSON（VBA 侧）。

用法：
    ./.venv/Scripts/python.exe diff_build_inputs.py

场景：
    TC21/TC24/TC20 —— 从 data/ 下已跑过 Python 版的 e2e 工作簿提取输入三表
    TC19  —— TC-19_E07 冻结数据（物流单号仅在库存表）
    TC36  —— TC-36_E11 冻结数据（QC 碎片库存）
    TCXB  —— 自编碎片场景（S<T<S+minQtyOther），用于实证已知偏差①（VBA 预检测B 漏检）

产物（均在 data/diff/ 下）：
    <case>_input.json  —— 输入三表数据，供 diff_run_vba.ps1 写入 VBA 测试工作簿
    <case>_py.xlsx     —— 写好输入、清空输出表/运行历史的工作簿，供 Python CLI allocate 原地跑
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
DIFF_DIR = ROOT / "data" / "diff"

# 以 e2e 工作簿为模板（含全部 8 张表与表头、文本格式列）
TEMPLATE = ROOT / "data" / "e2e_SF0060_TC20.xlsx"

# 从 e2e 工作簿提取输入的场景
E2E_CASES = {
    "TC21": ROOT / "data" / "e2e_SF0016_TC21.xlsx",  # 回溯成功场景
    "TC24": ROOT / "data" / "e2e_SF0028_TC24.xlsx",  # E10 回溯超限场景
    "TC20": ROOT / "data" / "e2e_SF0060_TC20.xlsx",  # E08 校验失败场景
}

# 冻结/自编输入的场景：returns/inventory 为数据行（不含表头），config 为 (参数名, 值)
LITERAL_CASES = {
    # TC-19 E07：物流单号仅存在于质检库存表（退单表空）
    "TC19": {
        "returns": [],
        "inventory": [["SF3190000000059", "H000000059", "ZP", "LA01", "2029/06/15", 5]],
        "config": [],
    },
    # TC-36 E11：ZP/QC 均为碎片（T=1 < groupMinQty=2）
    "TC36": {
        "returns": [["SF3190000000036", "TK00000036", "H000000001", "00001", 2]],
        "inventory": [
            ["SF3190000000036", "H000000001", "ZP", "LA01", "2029/01/01", 1],
            ["SF3190000000036", "H000000001", "QC", "LA01", "2029/01/01", 1],
        ],
        "config": [["调试日志级别", "简版"]],
    },
    # 自编碎片场景：S=20（00001/00002 锁定 ZP），T=21，minQtyOther=2，
    # 满足 S<T<S+minQtyOther → Python 预检测B 命中，VBA 漏检（已知偏差①实证）。
    # E08 守恒：退单合计 26 = ZP 21 + QC 5。
    "TCXB": {
        "returns": [
            ["SF3190000000090", "TK10000900", "H000000090", "00001", 10],
            ["SF3190000000090", "TK10000900", "H000000090", "00002", 10],
            ["SF3190000000090", "TK10000900", "H000000090", "00003", 2],
            ["SF3190000000090", "TK10000900", "H000000090", "00004", 2],
            ["SF3190000000090", "TK10000900", "H000000090", "00005", 2],
        ],
        "inventory": [
            ["SF3190000000090", "H000000090", "ZP", "LA01", "2029/06/15", 21],
            ["SF3190000000090", "H000000090", "QC", "LA01", "2029/06/15", 5],
        ],
        "config": [["调试日志级别", "简版"]],
    },
}

OUTPUT_SHEETS = ("分配状态汇总表", "成功分配明细表", "数据异常明细表", "调试日志", "运行历史记录表")
DEBUG_SHEET = "调试日志"


def sheet_data_rows(ws) -> list[list]:
    """读取第 2 行起的数据行（跳过全空行）。"""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(list(row))
    return rows


def clear_data(ws) -> None:
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def clear_inputs_and_outputs(wb) -> None:
    for name in ("输入_退单表", "输入_质检库存表", "输入_配置", *OUTPUT_SHEETS):
        clear_data(wb[name])
    for name in list(wb.sheetnames):
        if name.startswith(DEBUG_SHEET + "_") and name[len(DEBUG_SHEET) + 1 :].isdigit():
            del wb[name]


def write_rows(ws, rows: list[list]) -> None:
    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)


def emit(case: str, returns: list[list], inventory: list[list], config: list[list], wb) -> None:
    payload = {"case": case, "returns": returns, "inventory": inventory, "config": config}
    (DIFF_DIR / f"{case}_input.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    wb.save(DIFF_DIR / f"{case}_py.xlsx")
    print(f"{case}: 退单行={len(returns)} 库存行={len(inventory)} 配置={config}")


def main() -> None:
    DIFF_DIR.mkdir(parents=True, exist_ok=True)

    for case, src in E2E_CASES.items():
        wb = openpyxl.load_workbook(src)
        returns = sheet_data_rows(wb["输入_退单表"])
        inventory = sheet_data_rows(wb["输入_质检库存表"])
        config = [[r[0], r[1]] for r in sheet_data_rows(wb["输入_配置"])]
        clear_inputs_and_outputs(wb)
        write_rows(wb["输入_退单表"], returns)
        write_rows(wb["输入_质检库存表"], inventory)
        write_rows(wb["输入_配置"], config)
        emit(case, returns, inventory, config, wb)

    for case, data in LITERAL_CASES.items():
        wb = openpyxl.load_workbook(TEMPLATE)
        clear_inputs_and_outputs(wb)
        write_rows(wb["输入_退单表"], data["returns"])
        write_rows(wb["输入_质检库存表"], data["inventory"])
        write_rows(wb["输入_配置"], data["config"])
        emit(case, data["returns"], data["inventory"], data["config"], wb)


if __name__ == "__main__":
    main()
